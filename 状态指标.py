import pandas as pd
from openpyxl import load_workbook
from datetime import date,timedelta,datetime
# 读取Excel文件中的Activity工作表
df = pd.read_excel('myLife.xlsx', sheet_name='Activity')
last_day = df['日期'].iloc[-1]  # 获取最后一天的日期

# 下面代码如果修改了mylife中的数据，获取的日期为datetime类型，如果没有修改，获取的日期为str类型
# 这是修改mylife后使用的代码
now_day = datetime.strptime(last_day, '%Y-%m-%d').date() # 将字符串转换为日期
yesterday = now_day - timedelta(days=1) # 获取前一天的日期
yesterday = str(yesterday)  # 将前一天的日期转换为字符串类型
#这是没有修改mylife使用的代码
# yesterday = last_day - timedelta(days=1)

last_day_operations = df.loc[df['日期'] == last_day, '操作']    # 获取最后一天的操作内容
yesterday_operations = df.loc[df['日期'] == yesterday, '操作']    # 获取昨天的操作内容
total_operations = len(last_day_operations) # 统计操作内容的总数
health_score = 0    # 定义健康值变量
# 判断操作总数，并赋予相应的健康值,计算步行、喝水、睡眠、吃饭对健康值的贡献，以及类似,假设小于1000ml贡献0分，大于2000贡献10分
if total_operations >= 20:
    health_score = 60
else:
    health_score = 60 - 3 * (20 - total_operations)
print("基础健康值：",health_score)
# print("最后一天的日期：", last_day)
# print("操作内容的总数：", total_operations)
# 获取最新一天操作内容为"喝水"、“步行”的字段和对应的数量
last_day_water = df.loc[(df['日期'] == last_day) & (df['操作'] == '喝水'), '数量']
last_day_walk = df.loc[(df['日期'] == last_day) & (df['操作'] == '步行'), '数量']
# 获取最新一天操作内容为"吃饭"、“睡觉”的字段的时长
last_day_sleep = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['睡眠', '午睡'])), '时长'].sum()
last_day_meals = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['吃早餐', '吃午餐', '吃晚餐'])), '操作'].count()
last_day_entertainment = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['购物', '欣赏', '沏茶', '修养', '拍摄', '阅读','看电视','看电影',"运动"])), '操作'].count()
last_day_konwledge = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['编程', '预习', '阅读', '自学', '讨论', '实验', '写作', '思考', '上课'])), '时长'].sum()
last_day_play = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['玩手机','看电视','看电影',])), '时长'].sum()
last_day_focus = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['编程', '阅读', '自学','实验', '写作','上课'])), '时长'].sum()
print("吃饭次数：",last_day_meals)
print("知识增量时长：",last_day_konwledge)
print("娱乐次数：",last_day_entertainment)
print("专注时长：",last_day_focus)
# 通用函数  
# 该函数包含4个区间递增-递增-递减-直线
def calculate_score(x, a, b, c, score):     # x为从myLife中获取的数量或者时长，score为给予改操作的最大分数
    if x < a:
        score = ((x - a) / a) * score      # 在0到a的范围内将score从-score递增到0。
    elif a <= x <= b:
        score = (x - a) / (b - a) * score   # 在a到b的范围内将score从-score递增到score
    elif b < x <= c:
        score = (c - x) / (c - b) * score   # 在b到c的范围内将score从score递减到-score
    else:
        score = -score                      # 如果x大于c，则score被设置为-score
    return round(score, 1)
# 通用函数  
# 该函数包含3个区间递增-递增-直线
def calculate_score2(x, a, b, score):     # x为从myLife中获取的数量或者时长，score为给予改操作的最大分数
    if x < a:
        score = ((x - a) / a) * score      # 在0到a的范围内将score从-score递增到0。
    elif a <= x <= b:
        score = (x - a) / (b - a) * score   # 在a到b的范围内将score从0递增到score
    else:
        score = score                      # 如果x大于c，则score被设置为score
    return round(score, 1)
#定义吃饭健康值的计算函数
def meals_score():
    meals_score = 0
    if last_day_meals != 0:
        meals_count = last_day_meals
        if meals_count == 3:    #当meals_count为3时，赋予10分
            meals_score = 10.0
        elif meals_count == 2:    #当meals_count为2时，赋予8分
           meals_score = 8.0
        elif meals_count == 1:    #当meals_count为1时，赋予0分
            meals_score = 0.0
        elif meals_count == 0:    #当meals_count为0时，赋予-12分
            meals_score = -12.0
    return meals_score
meals = meals_score()           
print("吃饭健康值：", meals)
drink = calculate_score(x=last_day_water.iloc[0], a=1000, b=2100, c=3500, score=10)
print("喝水健康值：",drink)
walk = calculate_score2(x=last_day_walk.iloc[0], a=5000, b=10000, score=10)
print("步行健康值：",walk)
sleep = calculate_score(x=last_day_sleep, a=390, b=510, c=750, score=10)
print("睡眠健康值：",sleep)
health_score = round(health_score + meals + drink + walk + sleep,0) # 计算健康值最终分数
print("最终健康值：",health_score)

# 计算知识增量
konwledge_score = round(calculate_score(x=last_day_konwledge, a=20, b=480, c=1440, score=100),0)
print("知识增长量：",konwledge_score)

#计算情绪值
# 定义情绪值基础分
emo_score = 0    # 定义情绪值变量
if total_operations >= 20:
    emo_score = 60
else:
    emo_score = 60 - 3 * (20 - total_operations)
print("基础健康值：",emo_score)

if last_day_entertainment:      # 判断有关娱乐的操作，存在一个加1分，最多加10分
    entertainment = 2*last_day_entertainment
    if last_day_entertainment > 10.0:
        entertainment = 20.0
print("娱乐情绪值：",entertainment)
play = calculate_score(x=last_day_play, a=0, b=220, c=720, score=10)
print("玩手机情绪值：",play)
# 对比昨天是否有新增操作，有一个则加2分，上限20分。1.找出上周中新增的操作   
new_operations = set(last_day_operations) - set(yesterday_operations)
new_operations_score = 0
if new_operations:
    new_operations_score = len(new_operations)
    if new_operations_score > 10:
        new_operations_score = 10.0
print("新增操作情绪值：", new_operations_score)
emo_score = round((emo_score + play + entertainment + new_operations_score),0)
print("最终情绪值：",emo_score)

# 计算专注度
# 定义专注度基础分
focus_score = 0    # 定义情绪值变量
if total_operations >= 20:
    focus_score = 60
else:
    focus_score = 60 - 2 * (20 - total_operations)
print("基础专注度：",focus_score)
focus = round(calculate_score(x=last_day_focus,a=90,b=480,c=1440,score=40),0)
print("专注度:",focus)
focus_score = int(focus_score + focus)
print("最终专注度:",focus_score)

# 计算体验度,体验度基础80，每新增一个操作加2分，如果专注度大于90，则每大2就在体验度扣1分
experience_score = 0    # 定义情绪值变量
if total_operations >= 20:
    experience_score = 80
else:
    experience_score = 80 - 4 * (20 - total_operations)
print("基础体验度：",experience_score)
# 如果专注度大于90，则每大2就在体验度扣1分,在80-90区间每小2体验度就加1分，其余为0分
if focus_score>=90:
    a = (focus_score-90)/2
elif 80<focus_score<90:
    a = (focus_score-90)/2
elif focus_score<80:
    a = 0
experience_score = experience_score + new_operations_score - a
print("最终体验度：",experience_score)

# 计算产出分，产出分为所以分的平均分
produce_score = round((health_score + emo_score + focus_score + experience_score)/4,0)
print("产出分",produce_score)
# 将内容写进指标sheet表
workbook = load_workbook('myLife.xlsx')# 打开Excel文件
worksheet = workbook['指标']# 选择"指标"工作表
values = [
    (last_day, '健康值', health_score),
    (last_day, '专注度', focus_score),
    (last_day, '体验度', experience_score),
    (last_day, '情绪值', emo_score),
    (last_day, '知识增量', konwledge_score),
    (last_day, '产出分', produce_score)
]

# Find the last row in the sheet
last_row = worksheet.max_row + 1

# Write the values using a loop
for i, (day, label, score) in enumerate(values):
    worksheet.cell(row=last_row+i, column=1, value=day)
    worksheet.cell(row=last_row+i, column=2, value=label)
    worksheet.cell(row=last_row+i, column=3, value=score)
# 保存Excel文件
workbook.save('myLife.xlsx')

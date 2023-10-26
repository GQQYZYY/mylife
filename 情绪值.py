import pandas as pd
from openpyxl import load_workbook
from datetime import date,datetime,timedelta
# 读取Excel文件中的Activity工作表
df = pd.read_excel('myLife.xlsx', sheet_name='Activity')
last_day = df['日期'].iloc[-1]  # 获取最后一天的日期
# 将字符串转换为日期
now_day = datetime.strptime(last_day, '%Y-%m-%d').date()
# 获取前一天的日期
yesterday = now_day - timedelta(days=1)
print(yesterday)
# 获取倒数第二天的日期  
last_day_operations = df.loc[df['日期'] == last_day, '操作']    # 获取最后一天的操作内容
new_day_operations = df.loc[df['日期'] == now_day, '操作']    # 获取最后一天的操作内容
total_operations = len(last_day_operations) # 统计操作内容的总数
# 将最后一天和倒数第二天操作列转化为列表  
last_day_operations = df.loc[df['日期'] == last_day, '操作'].tolist()  
now_day_operations = df.loc[df['日期'] == now_day, '操作'].tolist() 
emotional_score = 0    # 定义情绪值变量
# 判断操作总数，并赋予相应的情绪值
# 计算步行、、睡眠、吃饭对情绪值的贡献，以及类似
# 假设小于1000ml贡献0分，大于2000贡献10分
if total_operations >= 20:
    emotional_score = 60
else:
    emotional_score = 60 - 3 * (20 - total_operations)
print(emotional_score)
# print("最后一天的日期：", last_day)
# print("操作内容的总数：", total_operations)
# print("情绪值：", emotional_score)
 # 获取最新一天操作内容为"喝水"的字段和对应的数量
last_day_water = df.loc[(df['日期'] == last_day) & (df['操作'] == '喝水'), '数量']
last_day_walk = df.loc[(df['日期'] == last_day) & (df['操作'] == '步行'), '数量']
last_day_sleep = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['睡眠', '午睡'])), '时长'].sum()
last_day_meals = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['吃早餐', '吃午餐', '吃晚餐'])), '操作'].count()
last_day_entertainment = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['购物', '欣赏',"沏茶","看电影"])), '时长'].count()
last_day_play = df.loc[(df['日期'] == last_day) & (df['操作'].isin(['玩手机', '网聊'])), '时长'].sum()  
now_day_operations = df.loc[df['日期'] == now_day, '操作'].tolist()  
  
# 计算新增操作的数量  
new_operations = len(set(last_day_operations) - set(now_day_operations))  
  
print("新增操作的数量：", new_operations)
  

#定义喝水情绪值的计算函数
def water_score():
    water_emotional_score = 0
    if not last_day_water.empty:
        water_quantity = last_day_water.iloc[0]
        if 1000 <= water_quantity <= 2300:
            water_emotional_score = (water_quantity - 1000) // 130
        elif water_quantity < 1000:
            water_emotional_score = (water_quantity - 900) // 100 - 1
        elif water_quantity > 2500 and water_quantity <= 4500:
            water_emotional_score = 10 - (water_quantity - 2500) // 100
        elif water_quantity > 4500:
            water_emotional_score = -10.0
    return water_emotional_score
water = water_score()           
print("喝水情绪值：", water)
# #定义步行情绪值的计算函数
# def walk_score():
#     walk_score = 0
#     if not last_day_walk.empty:
#         walk_quantity = last_day_walk.iloc[0]
#         if 5000 <= walk_quantity <= 10000:
#             walk_score = (walk_quantity - 5000) // 625
#         elif walk_quantity < 5000:
#             walk_score = (walk_quantity - 4500) // 625 - 1
#         elif walk_quantity > 10000:
#             walk_score = 8.0
#     return walk_score
# walk = walk_score()           
# print("步行情绪值：", walk)
#定义睡眠情绪值的计算函数
def sleep_score():
    sleep_score = 0
    if last_day_sleep != 0:
        sleep_quantity = last_day_sleep
        if 390 <= sleep_quantity <= 510:
            sleep_score = 12.0
        elif sleep_quantity < 360:
           sleep_score = round((360 - sleep_quantity) / 36, 1)
        elif 510 < sleep_quantity <= 750:
            sleep_score = round(12 - (sleep_quantity - 510) / 24, 1)
        elif sleep_quantity > 750:
            sleep_score = -12.0
    return sleep_score
sleep  = sleep_score()           
print("睡眠情绪值：", sleep)
#定义吃饭情绪值的计算函数
def meals_score():
    meals_score = 0
    if last_day_meals != 0:
        meal_count = last_day_meals
        if meal_count == 3:
            meals_score = 10.0
        elif meal_count == 2:
           meals_score = 8.0
        elif meal_count == 1:
            meals_score = 0.0
        elif meal_count == 0:
            meals_score = -12.0
    return meals_score
meals = meals_score()           
print("吃饭情绪值：", meals)
#定义娱乐情绪值的计算函数
def play_score():
    play_score = 0
    if last_day_play != 0:
        play_quantity = last_day_play
        if  180<= play_quantity <= 220:
            play_score = 10
        elif play_quantity < 180:
           play_score = round((180 - play_quantity) / 18, 1)
        elif 220 < play_quantity <= 300:
            play_score = round(10 - (play_quantity - 220) / 8, 1)
        elif play_quantity > 300:
            play_score = -10.0
    return play_score
play = play_score()           
print("玩乐情绪值：",play )
#定义休息娱乐情绪值的计算函数
def entertainment_score():
    entertainment_score = 0
    if last_day_meals != 0:
        entertainment_count = last_day_meals
        if entertainment_count == 4:
           entertainment_score = 12.0
        elif entertainment_count == 3:
           entertainment_score = 9.0
        elif entertainment_count == 2:
            entertainment_score = 6.0
        elif entertainment_count == 1:
            entertainment_score = 3.0
    return entertainment_score
entertainment = entertainment_score()           
print("吃饭情绪值：", entertainment)
emotional_score =emotional_score + meals + sleep  + play + entertainment#计算情绪值 
print("情绪值：", emotional_score)
# 将内容写进指标sheet表
workbook = load_workbook('myLife.xlsx')# 打开Excel文件
worksheet = workbook['指标']# 选择"指标"工作表
last_row = worksheet.max_row + 1# 获取最后一行的行号
# 将日期和健康分数写入Excel文件
worksheet.cell(row=last_row, column=1, value=last_day)
worksheet.cell(row=last_row, column=2, value='情绪值')
worksheet.cell(row=last_row, column=3, value=emotional_score)
# 保存Excel文件
workbook.save('myLife.xlsx') 